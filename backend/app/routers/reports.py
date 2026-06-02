from fastapi import APIRouter, Depends, Query, Response
from sqlalchemy.orm import Session
from sqlalchemy import func, cast, Date
from typing import Optional
from datetime import date, datetime, timedelta
import io

from ..database import get_db
from ..models import (Checkin, CheckinStatus, Invoice, Customer, Room,
                      RoomStatus, Alert, Setting)
from ..auth import get_current_user, require_admin, resolve_lodge_scope

router = APIRouter(prefix="/api/reports", tags=["reports"])


@router.get("/summary")
def get_summary_report(from_date: Optional[str] = None, to_date: Optional[str] = None,
                       db: Session = Depends(get_db),
                       current_user=Depends(get_current_user),
                       lodge_id: int = Depends(resolve_lodge_scope)):
    # Defaults to last 30 days if not provided
    if not from_date:
        from_dt = datetime.now() - timedelta(days=30)
    else:
        from_dt = datetime.fromisoformat(from_date)
    
    if not to_date:
        to_dt = datetime.now()
    else:
        to_dt = datetime.fromisoformat(to_date + "T23:59:59")

    # Metrics — all scoped to this lodge.
    total_revenue = db.query(func.sum(Invoice.total_amount)).filter(
        Invoice.lodge_id == lodge_id,
        Invoice.created_at >= from_dt, Invoice.created_at <= to_dt
    ).scalar() or 0
    
    checkins_count = db.query(Checkin).filter(
        Checkin.lodge_id == lodge_id,
        Checkin.checkin_datetime >= from_dt, Checkin.checkin_datetime <= to_dt
    ).count()
    
    total_guests = db.query(func.sum(Checkin.members_count)).filter(
        Checkin.lodge_id == lodge_id,
        Checkin.checkin_datetime >= from_dt, Checkin.checkin_datetime <= to_dt
    ).scalar() or 0
    
    new_customers = db.query(Customer).filter(
        Customer.lodge_id == lodge_id,
        Customer.created_at >= from_dt, Customer.created_at <= to_dt
    ).count()
    
    room_nights = db.query(func.sum(Checkin.total_nights)).filter(
        Checkin.lodge_id == lodge_id,
        Checkin.status == CheckinStatus.checked_out,
        Checkin.actual_checkout >= from_dt, Checkin.actual_checkout <= to_dt
    ).scalar() or 0
    
    total_rooms = db.query(Room).filter(
        Room.lodge_id == lodge_id, Room.is_active == True
    ).count()
    days = (to_dt - from_dt).days or 1
    max_room_nights = total_rooms * days
    avg_occupancy = (room_nights / max_room_nights * 100) if max_room_nights else 0
    
    avg_rev_per_night = (float(total_revenue) / room_nights) if room_nights else 0

    # Compute best room type by revenue in window (lodge-scoped)
    best_row = (db.query(Room.room_type, func.sum(Checkin.total_amount).label("rev"))
                .join(Checkin)
                .filter(Room.lodge_id == lodge_id,
                        Checkin.lodge_id == lodge_id,
                        Checkin.checkin_datetime >= from_dt,
                        Checkin.checkin_datetime <= to_dt,
                        Checkin.total_amount.isnot(None))
                .group_by(Room.room_type)
                .order_by(func.sum(Checkin.total_amount).desc())
                .first())
    rt_val = best_row[0].value if best_row and hasattr(best_row[0], "value") else (best_row[0] if best_row else None)
    best_room_type = (rt_val.replace("_", " ").upper() if rt_val else "—")

    return {
        "total_revenue": float(total_revenue),
        "checkins_count": checkins_count,
        "total_guests": int(total_guests),
        "new_customers": new_customers,
        "occupied_room_nights": int(room_nights),
        "avg_occupancy": round(avg_occupancy, 1),
        "avg_revenue_per_night": round(avg_rev_per_night, 2),
        "best_room_type": best_room_type
    }


@router.get("/room-types")
def get_room_type_stats(from_date: Optional[str] = None, to_date: Optional[str] = None,
                        db: Session = Depends(get_db),
                        current_user=Depends(get_current_user),
                        lodge_id: int = Depends(resolve_lodge_scope)):
    if not from_date: from_dt = datetime.now() - timedelta(days=30)
    else: from_dt = datetime.fromisoformat(from_date)
    if not to_date: to_dt = datetime.now()
    else: to_dt = datetime.fromisoformat(to_date + "T23:59:59")

    stats = db.query(
        Room.room_type,
        func.count(Checkin.checkin_id).label("stays"),
        func.sum(Checkin.total_nights).label("nights"),
        func.sum(Checkin.total_amount).label("revenue")
    ).join(Checkin).filter(
        Room.lodge_id == lodge_id,
        Checkin.lodge_id == lodge_id,
        Checkin.checkin_datetime >= from_dt,
        Checkin.checkin_datetime <= to_dt
    ).group_by(Room.room_type).all()

    total_revenue = sum(float(s.revenue or 0) for s in stats)
    
    return [
        {
            "room_type": (s.room_type.value if hasattr(s.room_type, "value") else s.room_type).replace("_", " ").upper(),
            "stays": s.stays,
            "nights": int(s.nights or 0),
            "revenue": float(s.revenue or 0),
            "pct": round((float(s.revenue or 0) / total_revenue * 100) if total_revenue else 0, 1),
            "avg_per_night": round((float(s.revenue or 0) / s.nights) if s.nights else 0, 2)
        }
        for s in stats
    ]


@router.get("/dashboard")
def get_dashboard_data(db: Session = Depends(get_db),
                       current_user=Depends(get_current_user),
                       lodge_id: int = Depends(resolve_lodge_scope)):
    today = date.today()
    now = datetime.now()

    total_rooms = db.query(Room).filter(
        Room.lodge_id == lodge_id, Room.is_active == True
    ).count()
    available_rooms = db.query(Room).filter(
        Room.lodge_id == lodge_id,
        Room.status == RoomStatus.available, Room.is_active == True
    ).count()
    occupied_rooms = db.query(Checkin).filter(
        Checkin.lodge_id == lodge_id,
        Checkin.status == CheckinStatus.active
    ).count()
    blocked_rooms = db.query(Room).filter(
        Room.lodge_id == lodge_id,
        Room.status == RoomStatus.blocked, Room.is_active == True
    ).count()
    maintenance_rooms = db.query(Room).filter(
        Room.lodge_id == lodge_id,
        Room.status == RoomStatus.maintenance, Room.is_active == True
    ).count()
    today_start = datetime.combine(today, datetime.min.time())
    today_end = datetime.combine(today, datetime.max.time())
    due_checkout = db.query(Checkin).filter(
        Checkin.lodge_id == lodge_id,
        Checkin.status == CheckinStatus.active,
        Checkin.expected_checkout >= today_start,
        Checkin.expected_checkout <= today_end,
    ).count()
    total_customers = db.query(Customer).filter(
        Customer.lodge_id == lodge_id, Customer.is_active == True
    ).count()
    today_revenue = db.query(func.sum(Invoice.total_amount)).filter(
        Invoice.lodge_id == lodge_id,
        cast(Invoice.created_at, Date) == today
    ).scalar() or 0
    overdue_count = db.query(Checkin).filter(
        Checkin.lodge_id == lodge_id,
        Checkin.status == CheckinStatus.active,
        Checkin.expected_checkout < now,
        Checkin.expected_checkout.isnot(None)
    ).count()

    recent_checkins = db.query(Checkin).filter(
        Checkin.lodge_id == lodge_id,
        cast(Checkin.checkin_datetime, Date) == today
    ).order_by(Checkin.checkin_datetime.desc()).limit(5).all()
    recent_checkouts = db.query(Checkin).filter(
        Checkin.lodge_id == lodge_id,
        Checkin.status == CheckinStatus.checked_out,
        cast(Checkin.actual_checkout, Date) == today
    ).order_by(Checkin.actual_checkout.desc()).limit(5).all()

    activity = []
    for ch in recent_checkins:
        if not ch.customer or not ch.room:
            continue
        activity.append({
            "type": "checkin", "icon": "🟢",
            "message": f"{ch.customer.first_name} {ch.customer.last_name} checked into Room {ch.room.room_number}",
            "time": ch.checkin_datetime.isoformat() if ch.checkin_datetime else ""
        })
    for ch in recent_checkouts:
        if not ch.customer or not ch.room or not ch.actual_checkout:
            continue
        activity.append({
            "type": "checkout", "icon": "🔴",
            "message": f"{ch.customer.first_name} {ch.customer.last_name} checked out of Room {ch.room.room_number}",
            "time": ch.actual_checkout.isoformat()
        })
    activity.sort(key=lambda x: x["time"], reverse=True)

    from sqlalchemy import case
    room_breakdown = db.query(
        Room.room_type,
        func.count(Room.room_id).label("total"),
        func.sum(case((Room.status == RoomStatus.available, 1), else_=0)).label("available"),
        func.sum(case((Room.status == RoomStatus.occupied, 1), else_=0)).label("occupied"),
        func.sum(case((Room.status == RoomStatus.blocked, 1), else_=0)).label("blocked"),
        func.sum(case((Room.status == RoomStatus.maintenance, 1), else_=0)).label("maintenance")
    ).filter(
        Room.lodge_id == lodge_id, Room.is_active == True
    ).group_by(Room.room_type).all()

    chart_data = [{"room_type": r.room_type.value if hasattr(r.room_type, "value") else r.room_type,
                   "total": r.total,
                   "available": int(r.available or 0),
                   "occupied": int(r.occupied or 0),
                   "blocked": int(r.blocked or 0),
                   "maintenance": int(r.maintenance or 0)} for r in room_breakdown]

    thirty_days_ago = today - timedelta(days=30)
    daily_checkins = db.query(
        func.date(Checkin.checkin_datetime).label("day"),
        func.count(Checkin.checkin_id).label("count")
    ).filter(
        Checkin.lodge_id == lodge_id,
        Checkin.checkin_datetime >= thirty_days_ago
    ).group_by(func.date(Checkin.checkin_datetime)).order_by(func.date(Checkin.checkin_datetime)).all()

    return {
        "kpis": {
            "total_rooms": total_rooms,
            "available_rooms": available_rooms,
            "occupied_rooms": occupied_rooms,
            "blocked_rooms": blocked_rooms,
            "maintenance_rooms": maintenance_rooms,
            "due_checkout_today": due_checkout,
            "total_customers": total_customers,
            "today_revenue": float(today_revenue),
            "overdue_count": overdue_count,
            "occupancy_rate": round((occupied_rooms / total_rooms * 100) if total_rooms else 0, 1)
        },
        "activity": activity[:10],
        "room_breakdown": chart_data,
        "daily_checkins": [{"day": str(d.day), "count": d.count} for d in daily_checkins]
    }


@router.get("/occupancy")
def occupancy_report(from_date: Optional[str] = None, to_date: Optional[str] = None,
                     room_type: Optional[str] = None,
                     db: Session = Depends(get_db),
                     current_user=Depends(get_current_user),
                     lodge_id: int = Depends(resolve_lodge_scope)):
    query = db.query(Checkin).filter(Checkin.lodge_id == lodge_id)
    if from_date:
        query = query.filter(Checkin.checkin_datetime >= datetime.fromisoformat(from_date))
    if to_date:
        query = query.filter(Checkin.checkin_datetime <= datetime.fromisoformat(to_date + "T23:59:59"))
    if room_type:
        query = query.join(Room).filter(Room.lodge_id == lodge_id, Room.room_type == room_type)

    checkins = query.all()
    total_checkins = len(checkins)
    total_nights = sum(ch.total_nights or 0 for ch in checkins)
    avg_stay = total_nights / total_checkins if total_checkins else 0

    if not from_date: from_dt = datetime.now() - timedelta(days=30)
    else: from_dt = datetime.fromisoformat(from_date)
    if not to_date: to_dt = datetime.now()
    else: to_dt = datetime.fromisoformat(to_date + "T23:59:59")

    chart_data = []
    curr = from_dt
    while curr <= to_dt:
        d_str = curr.strftime("%Y-%m-%d")
        count = sum(1 for ch in checkins if ch.checkin_datetime.date() == curr.date())
        total_rooms = db.query(Room).filter(Room.lodge_id == lodge_id).count() or 1
        chart_data.append({
            "date": d_str,
            "occupancy_pct": round((count / total_rooms * 100), 1)
        })
        curr += timedelta(days=1)

    return chart_data


@router.get("/revenue")
def revenue_report(from_date: Optional[str] = None, to_date: Optional[str] = None,
                   db: Session = Depends(get_db),
                   current_user=Depends(get_current_user),
                   lodge_id: int = Depends(resolve_lodge_scope)):
    query = db.query(Invoice).filter(Invoice.lodge_id == lodge_id)
    if from_date:
        query = query.filter(Invoice.created_at >= datetime.fromisoformat(from_date))
    if to_date:
        query = query.filter(Invoice.created_at <= datetime.fromisoformat(to_date + "T23:59:59"))

    invoices = query.all()
    total_revenue = sum(float(inv.total_amount) for inv in invoices)
    total_gst = sum(float(inv.gst_amount or 0) for inv in invoices)
    total_discount = sum(float(inv.discount or 0) for inv in invoices)

    by_payment = {}
    for inv in invoices:
        pm = inv.payment_mode or "cash"
        by_payment.setdefault(pm, {"count": 0, "amount": 0})
        by_payment[pm]["count"] += 1
        by_payment[pm]["amount"] += float(inv.total_amount)

    by_day = {}
    for inv in invoices:
        day = inv.created_at.strftime("%Y-%m-%d") if inv.created_at else "unknown"
        by_day.setdefault(day, 0)
        by_day[day] += float(inv.total_amount)

    return [{"date": d, "revenue": v} for d, v in sorted(by_day.items())]


@router.get("/outstanding")
def outstanding_dues(db: Session = Depends(get_db),
                     current_user=Depends(get_current_user),
                     lodge_id: int = Depends(resolve_lodge_scope)):
    now = datetime.now()
    today = now.date()
    overdue = db.query(Checkin).filter(
        Checkin.lodge_id == lodge_id,
        Checkin.status == CheckinStatus.active,
        Checkin.expected_checkout < now,
        Checkin.expected_checkout.isnot(None)
    ).all()

    result = []
    for ch in overdue:
        # expected_checkout is a DateTime — take .date() before subtracting
        # dates so we get whole-day overdue counts.
        days_overdue = (today - ch.expected_checkout.date()).days
        result.append({
            "checkin_id": ch.checkin_id,
            "customer_name": f"{ch.customer.first_name} {ch.customer.last_name}",
            "customer_phone": ch.customer.phone,
            "room_number": ch.room.room_number,
            "checkin_date": ch.checkin_datetime.isoformat(),
            "expected_checkout": ch.expected_checkout.isoformat(),
            "days_overdue": days_overdue,
            "deposit_amount": float(ch.deposit_amount),
            "estimated_charges": days_overdue * float(ch.tariff_per_night)
        })
    return result


@router.get("/export")
def export_report(
    report: str = Query(...),
    format: str = Query("xlsx"),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    lodge_id: int = Depends(resolve_lodge_scope),
):
    """Export reports as Excel or PDF — scoped to the current lodge."""
    if format == "pdf":
        return generate_pdf_report(report, from_date, to_date, db, lodge_id)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B2A4A")

    if report == "checkins":
        ws.title = "Check-in History"
        headers = ["Checkin ID", "Customer", "Phone", "Room", "Room Type",
                   "Check-in", "Checkout", "Nights", "Amount", "Status"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        q = db.query(Checkin).filter(Checkin.lodge_id == lodge_id)
        if from_date:
            q = q.filter(Checkin.checkin_datetime >= datetime.fromisoformat(from_date))
        if to_date:
            q = q.filter(Checkin.checkin_datetime <= datetime.fromisoformat(to_date + "T23:59:59"))

        for row_idx, ch in enumerate(q.order_by(Checkin.checkin_datetime.desc()).all(), 2):
            ws.cell(row=row_idx, column=1, value=ch.checkin_id)
            ws.cell(row=row_idx, column=2, value=f"{ch.customer.first_name} {ch.customer.last_name}" if ch.customer else "")
            ws.cell(row=row_idx, column=3, value=ch.customer.phone if ch.customer else "")
            ws.cell(row=row_idx, column=4, value=ch.room.room_number if ch.room else "")
            ws.cell(row=row_idx, column=5, value=ch.room.room_type if ch.room else "")
            ws.cell(row=row_idx, column=6, value=ch.checkin_datetime.strftime("%d/%m/%Y %H:%M") if ch.checkin_datetime else "")
            ws.cell(row=row_idx, column=7, value=ch.actual_checkout.strftime("%d/%m/%Y %H:%M") if ch.actual_checkout else "")
            ws.cell(row=row_idx, column=8, value=ch.total_nights or "")
            ws.cell(row=row_idx, column=9, value=float(ch.total_amount) if ch.total_amount else "")
            ws.cell(row=row_idx, column=10, value=ch.status)

    elif report == "revenue":
        ws.title = "Revenue Report"
        headers = ["Invoice No", "Customer", "Room", "Checkout Date", "Nights",
                   "Room Charges", "GST", "Discount", "Total", "Payment Mode"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        q = db.query(Invoice).filter(Invoice.lodge_id == lodge_id)
        if from_date:
            q = q.filter(Invoice.created_at >= datetime.fromisoformat(from_date))
        if to_date:
            q = q.filter(Invoice.created_at <= datetime.fromisoformat(to_date + "T23:59:59"))

        for row_idx, inv in enumerate(q.order_by(Invoice.created_at.desc()).all(), 2):
            ws.cell(row=row_idx, column=1, value=inv.invoice_number)
            ws.cell(row=row_idx, column=2, value=f"{inv.customer.first_name} {inv.customer.last_name}" if inv.customer else "")
            ws.cell(row=row_idx, column=3, value=inv.room.room_number if inv.room else "")
            ws.cell(row=row_idx, column=4, value=inv.checkout_datetime.strftime("%d/%m/%Y") if inv.checkout_datetime else "")
            ws.cell(row=row_idx, column=5, value=inv.nights)
            ws.cell(row=row_idx, column=6, value=float(inv.room_charges))
            ws.cell(row=row_idx, column=7, value=float(inv.gst_amount or 0))
            ws.cell(row=row_idx, column=8, value=float(inv.discount or 0))
            ws.cell(row=row_idx, column=9, value=float(inv.total_amount))
            ws.cell(row=row_idx, column=10, value=inv.payment_mode or "")

    elif report == "customers":
        ws.title = "Customers"
        headers = ["ID", "First Name", "Last Name", "Phone", "Email", "ID Type",
                   "Nationality", "Total Visits", "VIP", "Registered"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        customers = (db.query(Customer)
                     .filter(Customer.lodge_id == lodge_id, Customer.is_active == True)
                     .all())
        for row_idx, c in enumerate(customers, 2):
            ws.cell(row=row_idx, column=1, value=c.customer_id)
            ws.cell(row=row_idx, column=2, value=c.first_name)
            ws.cell(row=row_idx, column=3, value=c.last_name)
            ws.cell(row=row_idx, column=4, value=c.phone)
            ws.cell(row=row_idx, column=5, value=c.email or "")
            ws.cell(row=row_idx, column=6, value=c.id_type)
            ws.cell(row=row_idx, column=7, value=c.nationality or "Indian")
            ws.cell(row=row_idx, column=8, value=c.total_visits or 0)
            ws.cell(row=row_idx, column=9, value="Yes" if c.is_vip else "No")
            ws.cell(row=row_idx, column=10, value=c.created_at.strftime("%d/%m/%Y") if c.created_at else "")

    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 30)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"{report}_report_{date.today().isoformat()}.xlsx"
    return Response(
        content=stream.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def generate_pdf_report(report: str, from_date: Optional[str], to_date: Optional[str],
                         db: Session, lodge_id: int):
    """Helper to generate PDF reports using reportlab. All queries scoped to lodge."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    import io

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=10*mm, leftMargin=10*mm,
                             topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    story = []

    # Hotel header from THIS lodge's settings
    hotel_name = db.query(Setting).filter(
        Setting.setting_key == "hotel_name",
        Setting.lodge_id == lodge_id,
    ).first()
    hotel_name = hotel_name.setting_value if hotel_name else "Lodge"
    
    story.append(Paragraph(f"<b>{hotel_name}</b>", ParagraphStyle("title", fontSize=18, textColor=colors.HexColor("#1B2A4A"))))
    
    report_title = report.replace("_", " ").title() + " Report"
    date_range = f"({from_date or 'Start'} to {to_date or 'Today'})"
    story.append(Paragraph(f"{report_title} {date_range}", styles["Heading2"]))
    story.append(Spacer(1, 5*mm))

    data = []
    col_widths = []

    if report == "checkins":
        data.append(["ID", "Customer", "Phone", "Room", "Type", "Check-in", "Checkout", "Nights", "Amount", "Status"])
        col_widths = [15*mm, 45*mm, 30*mm, 15*mm, 25*mm, 35*mm, 35*mm, 15*mm, 25*mm, 25*mm]
        
        q = db.query(Checkin).filter(Checkin.lodge_id == lodge_id)
        if from_date:
            q = q.filter(Checkin.checkin_datetime >= datetime.fromisoformat(from_date))
        if to_date:
            q = q.filter(Checkin.checkin_datetime <= datetime.fromisoformat(to_date + "T23:59:59"))
        
        for ch in q.order_by(Checkin.checkin_datetime.desc()).all():
            data.append([
                str(ch.checkin_id),
                f"{ch.customer.first_name} {ch.customer.last_name}" if ch.customer else "",
                ch.customer.phone if ch.customer else "",
                ch.room.room_number if ch.room else "",
                ch.room.room_type if ch.room else "",
                ch.checkin_datetime.strftime("%d/%m/%y %H:%M") if ch.checkin_datetime else "",
                ch.actual_checkout.strftime("%d/%m/%y %H:%M") if ch.actual_checkout else "",
                str(ch.total_nights or ""),
                f"{float(ch.total_amount):.2f}" if ch.total_amount else "",
                ch.status
            ])

    elif report == "revenue":
        data.append(["Invoice No", "Customer", "Room", "Date", "Nights", "Charges", "GST", "Disc", "Total", "Mode"])
        col_widths = [35*mm, 45*mm, 20*mm, 30*mm, 15*mm, 25*mm, 25*mm, 25*mm, 25*mm, 25*mm]
        
        q = db.query(Invoice).filter(Invoice.lodge_id == lodge_id)
        if from_date:
            q = q.filter(Invoice.created_at >= datetime.fromisoformat(from_date))
        if to_date:
            q = q.filter(Invoice.created_at <= datetime.fromisoformat(to_date + "T23:59:59"))
            
        for inv in q.order_by(Invoice.created_at.desc()).all():
            data.append([
                inv.invoice_number,
                f"{inv.customer.first_name} {inv.customer.last_name}" if inv.customer else "",
                inv.room.room_number if inv.room else "",
                inv.checkout_datetime.strftime("%d/%m/%Y") if inv.checkout_datetime else "",
                str(inv.nights),
                f"{float(inv.room_charges):.2f}",
                f"{float(inv.gst_amount or 0):.2f}",
                f"{float(inv.discount or 0):.2f}",
                f"{float(inv.total_amount):.2f}",
                inv.payment_mode or ""
            ])

    elif report == "customers":
        data.append(["ID", "Name", "Phone", "Email", "ID Type", "Nationality", "Visits", "VIP", "Registered"])
        col_widths = [15*mm, 45*mm, 30*mm, 45*mm, 30*mm, 30*mm, 15*mm, 15*mm, 30*mm]
        
        customers = (db.query(Customer)
                     .filter(Customer.lodge_id == lodge_id, Customer.is_active == True)
                     .all())
        for c in customers:
            data.append([
                str(c.customer_id),
                f"{c.first_name} {c.last_name}",
                c.phone,
                c.email or "",
                c.id_type,
                c.nationality or "Indian",
                str(c.total_visits or 0),
                "Yes" if c.is_vip else "No",
                c.created_at.strftime("%d/%m/%Y") if c.created_at else ""
            ])

    if not data:
        story.append(Paragraph("No data found for the selected criteria.", styles["Normal"]))
    else:
        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B2A4A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ]))
        story.append(tbl)

    doc.build(story)
    buf.seek(0)
    
    filename = f"{report}_report_{date.today().isoformat()}.pdf"
    return Response(
        content=buf.read(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ════════════════════════════════════════════════════════════════════
#  Industry-standard PMS KPIs
# ════════════════════════════════════════════════════════════════════
# Reference (definitions used by every modern PMS):
#   ARR   = Average Room Rate           = total room revenue / rooms sold
#   ADR   = Average Daily Rate          = same as ARR (synonym)
#   OCC   = Occupancy %                 = rooms sold / room-nights available * 100
#   RevPAR = Revenue Per Available Room = room_revenue / room-nights available
#                                       = ARR * (Occupancy% / 100)
#   ALOS  = Average Length Of Stay      = total nights / number of checkouts
# These are the standard metrics every PMS dashboard exposes.

@router.get("/kpis")
def kpis(from_date: Optional[str] = None, to_date: Optional[str] = None,
         db: Session = Depends(get_db),
         current_user=Depends(get_current_user),
         lodge_id: int = Depends(resolve_lodge_scope)):
    """Hotel-industry KPIs for a date range: ARR, RevPAR, Occupancy, ALOS,
    plus revenue and expense totals so profit can be derived in one call."""
    today = date.today()
    fd = date.fromisoformat(from_date) if from_date else today - timedelta(days=29)
    td = date.fromisoformat(to_date) if to_date else today
    if td < fd:
        fd, td = td, fd  # tolerate reversed inputs

    days = (td - fd).days + 1

    # Count of total physical rooms in this lodge — used as denominator
    # for occupancy and RevPAR. We don't divide by zero even if the lodge
    # has no rooms yet.
    total_rooms = (db.query(func.count(Room.room_id))
                   .filter(Room.lodge_id == lodge_id,
                           Room.is_active == True).scalar()) or 0
    available_room_nights = total_rooms * days

    # "Sold" = nights where a check-in occupied a room within the window.
    # We approximate with checkouts whose stay falls in the window. This
    # matches the existing summary endpoint's logic so numbers tie out.
    sold_q = (db.query(Checkin)
              .filter(Checkin.lodge_id == lodge_id,
                      Checkin.status == CheckinStatus.checked_out,
                      cast(Checkin.actual_checkout, Date) >= fd,
                      cast(Checkin.actual_checkout, Date) <= td))
    rooms_sold = sold_q.count()
    nights_sold = (db.query(func.coalesce(func.sum(Checkin.total_nights), 0))
                   .filter(Checkin.lodge_id == lodge_id,
                           Checkin.status == CheckinStatus.checked_out,
                           cast(Checkin.actual_checkout, Date) >= fd,
                           cast(Checkin.actual_checkout, Date) <= td)
                   .scalar()) or 0

    # Room revenue: nights * tariff_per_night summed. We use the Checkin's
    # tariff so promotional rates / per-room overrides are respected.
    room_revenue = float((db.query(
        func.coalesce(func.sum(Checkin.total_nights * Checkin.tariff_per_night), 0))
        .filter(Checkin.lodge_id == lodge_id,
                Checkin.status == CheckinStatus.checked_out,
                cast(Checkin.actual_checkout, Date) >= fd,
                cast(Checkin.actual_checkout, Date) <= td)
        .scalar()) or 0)

    # Total invoiced revenue (includes extras + GST - discounts).
    total_revenue = float((db.query(func.coalesce(func.sum(Invoice.total_amount), 0))
                           .filter(Invoice.lodge_id == lodge_id,
                                   cast(Invoice.created_at, Date) >= fd,
                                   cast(Invoice.created_at, Date) <= td)
                           .scalar()) or 0)

    # Expenses for the same window — to derive net profit.
    expenses = 0.0
    try:
        from ..models import Expense
        expenses = float((db.query(func.coalesce(func.sum(Expense.amount), 0))
                          .filter(Expense.lodge_id == lodge_id,
                                  Expense.expense_date >= fd,
                                  Expense.expense_date <= td)
                          .scalar()) or 0)
    except Exception:
        # Expense table may not yet exist on a half-migrated DB
        pass

    # Compute KPIs. Guard every divisor.
    arr = (room_revenue / nights_sold) if nights_sold > 0 else 0.0
    occupancy_pct = (nights_sold / available_room_nights * 100) if available_room_nights > 0 else 0.0
    revpar = (room_revenue / available_room_nights) if available_room_nights > 0 else 0.0
    alos = (nights_sold / rooms_sold) if rooms_sold > 0 else 0.0

    return {
        "from": fd.isoformat(),
        "to": td.isoformat(),
        "days_in_range": days,
        "total_rooms": total_rooms,
        "available_room_nights": available_room_nights,
        "rooms_sold": rooms_sold,
        "nights_sold": int(nights_sold),
        "room_revenue": round(room_revenue, 2),
        "total_revenue": round(total_revenue, 2),
        "expenses": round(expenses, 2),
        "net_profit": round(total_revenue - expenses, 2),
        # Industry KPIs
        "arr": round(arr, 2),                  # Average Room Rate
        "occupancy_pct": round(occupancy_pct, 2),
        "revpar": round(revpar, 2),            # Revenue Per Available Room
        "alos": round(alos, 2),                # Average Length Of Stay
    }


@router.get("/forecast")
def forecast(days_ahead: int = 14,
             db: Session = Depends(get_db),
             current_user=Depends(get_current_user),
             lodge_id: int = Depends(resolve_lodge_scope)):
    """Forward-looking occupancy projection from confirmed bookings.
    Returns a day-by-day count of expected occupied rooms for the next
    `days_ahead` days. The Dashboard uses this to give staff visibility
    into the upcoming workload."""
    days_ahead = max(1, min(days_ahead, 90))
    from ..models import Booking, BookingStatus
    today = date.today()
    end = today + timedelta(days=days_ahead - 1)

    # Pull bookings whose date range overlaps the forecast window.
    rows = (db.query(Booking)
            .filter(Booking.lodge_id == lodge_id,
                    Booking.status.in_([BookingStatus.confirmed, BookingStatus.pending]),
                    Booking.checkin_date <= end,
                    Booking.checkout_date > today)
            .all())

    # Walk each booking and accumulate rooms-occupied per day. Bucketing
    # is more accurate than the simpler "count bookings" approach because
    # multi-night bookings legitimately occupy a room each night.
    by_day = {(today + timedelta(days=i)).isoformat(): 0
              for i in range(days_ahead)}
    for b in rows:
        cur = max(b.checkin_date, today)
        finish = min(b.checkout_date, end + timedelta(days=1))
        rooms = int(b.rooms_count or 1)
        while cur < finish:
            key = cur.isoformat()
            if key in by_day:
                by_day[key] += rooms
            cur += timedelta(days=1)

    total_rooms = (db.query(func.count(Room.room_id))
                   .filter(Room.lodge_id == lodge_id,
                           Room.is_active == True).scalar()) or 0
    series = []
    for d in sorted(by_day.keys()):
        occupied = by_day[d]
        pct = (occupied / total_rooms * 100) if total_rooms > 0 else 0
        series.append({
            "date": d,
            "occupied": occupied,
            "total_rooms": total_rooms,
            "occupancy_pct": round(pct, 1),
        })
    return {"days_ahead": days_ahead, "series": series}
