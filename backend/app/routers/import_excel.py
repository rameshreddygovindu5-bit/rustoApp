from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException
from sqlalchemy.orm import Session
from typing import Optional
from datetime import date, datetime
import io, re
from ..database import get_db
from ..models import Customer
from ..auth import get_current_user, resolve_lodge_scope
from ..services.alert_service import normalize_indian_phone

router = APIRouter(prefix="/api/import", tags=["import"])

# Header aliases — when the spreadsheet column matches any of these
# (case-insensitive, trimmed), we map it to the canonical Customer field.
# Adding a field here without an entry in the frontend's REQUIRED_FIELDS
# list means the column is auto-detected but not shown in the UI mapper.
COLUMN_ALIASES = {
    "first_name":    ["first name", "firstname", "fname", "given name", "first"],
    "last_name":     ["last name", "lastname", "lname", "surname", "family name", "last"],
    "phone":         ["phone", "mobile", "phone number", "contact", "cell", "mobile no",
                      "mobile number", "phone no"],
    "email":         ["email", "email address", "e-mail", "mail"],
    "address":       ["address", "full address", "residence", "addr"],
    "city":          ["city", "town"],
    "state":         ["state", "province"],
    "id_type":       ["id type", "document type", "id proof type", "id_type"],
    "id_number":     ["id number", "document number", "aadhar no", "aadhaar no",
                      "aadhar number", "dl no", "id_number", "id no"],
    "date_of_birth": ["dob", "date of birth", "birth date", "birthdate", "birth_date"],
    "gender":        ["gender", "sex"],
    "nationality":   ["nationality", "country"],
}


def fuzzy_match_column(header: str) -> Optional[str]:
    h = header.lower().strip()
    for field, aliases in COLUMN_ALIASES.items():
        if h in aliases or h == field:
            return field
    return None


@router.get("/template")
def download_template(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Generate and download a sample Excel template that exactly matches
    what /api/import/process accepts. Each column header here is one of the
    fuzzy-match aliases in COLUMN_ALIASES, so uploading the template back
    auto-maps without the user touching the column mapper.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi import Response

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

    headers = [
        "First Name", "Last Name", "Phone", "Email", "Address",
        "City", "State", "ID Type", "ID Number",
        "Date of Birth", "Gender", "Nationality",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B2A4A")
    header_align = Alignment(horizontal="left", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        # Width that fits both the header and typical Indian data.
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(h) + 4)

    # Two sample rows demonstrating the accepted formats.
    ws.append([
        "Ravi", "Kumar", "9876543210", "ravi.kumar@example.com",
        "Plot 21, MG Road", "Bengaluru", "Karnataka",
        "Aadhar", "123456789012",
        "1990-05-12", "Male", "Indian",
    ])
    ws.append([
        "Priya", "Sharma", "9123456780", "",
        "Flat 4B, Sector 22", "Hyderabad", "Telangana",
        "DL", "KA0120191234567",
        "1995-11-30", "Female", "Indian",
    ])

    ws.freeze_panes = "A2"

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return Response(
        content=stream.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customer_import_template.xlsx"}
    )


@router.post("/preview")
async def preview_excel(file: UploadFile = File(...),
                        db: Session = Depends(get_db),
                        current_user=Depends(get_current_user)):
    """Preview first 5 rows and suggest column mapping."""
    ext = file.filename.split(".")[-1].lower()
    content = await file.read()

    try:
        import openpyxl
        if ext in ["xlsx", "xls"]:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        elif ext == "csv":
            import csv
            text = content.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Use .xlsx, .xls, or .csv")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read file: {str(e)}")

    if not rows:
        raise HTTPException(status_code=400, detail="File is empty")

    headers = [str(h or "").strip() for h in rows[0]]
    preview_rows = [list(r) for r in rows[1:6]]

    # Auto-suggest mapping
    suggested_mapping = {}
    for idx, header in enumerate(headers):
        field = fuzzy_match_column(header)
        if field:
            suggested_mapping[field] = idx

    return {
        "headers": headers,
        "preview_rows": preview_rows,
        "suggested_mapping": suggested_mapping,
        "total_rows": len(rows) - 1,
    }


ID_TYPE_MAP = {
    "aadhar": "aadhar", "aadhaar": "aadhar", "aadhar card": "aadhar",
    "driving license": "driving_license", "driving licence": "driving_license",
    "dl": "driving_license", "driving_license": "driving_license",
    "voter": "voter_id", "voter id": "voter_id", "voter_id": "voter_id",
    "passport": "passport",
    "pan": "pan", "pan card": "pan",
}


def _parse_dob(value) -> Optional[date]:
    """Excel often delivers dates as datetime objects; CSV gives strings.
    Accept both, plus the common Indian DD-MM-YYYY and ISO YYYY-MM-DD forms."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unrecognized date format: '{s}'")


@router.post("/process")
async def process_import(
    file: UploadFile = File(...),
    mapping: str = Form(""),  # JSON string of {field: column_index}
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    lodge_id: int = Depends(resolve_lodge_scope),
):
    """Process and import customers from Excel/CSV.

    All imported rows are assigned to the current lodge. Duplicate detection
    is per-lodge (a phone that already exists in another lodge is fine — it
    represents a different business's record).
    """
    import json
    ext = file.filename.split(".")[-1].lower()
    content = await file.read()

    try:
        if ext in ["xlsx", "xls"]:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        elif ext == "csv":
            import csv
            text = content.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
        else:
            raise HTTPException(status_code=400, detail="Unsupported format")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Read error: {str(e)}")

    if not rows or len(rows) < 2:
        raise HTTPException(status_code=400, detail="File has no data rows.")

    try:
        col_map = json.loads(mapping) if mapping else {}
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Mapping is not valid JSON.")

    headers = [str(h or "").strip() for h in rows[0]]

    # Auto-detect if the frontend didn't supply a mapping.
    if not col_map:
        for idx, header in enumerate(headers):
            field = fuzzy_match_column(header)
            if field is not None:
                col_map[field] = idx

    # Drop entries with no real column index (the frontend can send -1 when
    # the user picked "— Not mapped —" or for an unknown column label).
    col_map = {k: int(v) for k, v in col_map.items()
               if v is not None and int(v) >= 0}

    missing_required = [f for f in ("first_name", "last_name", "phone")
                        if f not in col_map]
    if missing_required:
        pretty = {"first_name": "First Name",
                  "last_name": "Last Name",
                  "phone": "Phone"}
        raise HTTPException(
            status_code=400,
            detail=("Unmapped required column(s): "
                    f"{', '.join(pretty[k] for k in missing_required)}. "
                    f"Pick them in the mapping step.")
        )

    imported = 0
    duplicates = 0
    errors = []

    for row_num, raw_row in enumerate(rows[1:], 2):
        # Skip completely blank rows so users don't see hundreds of
        # spurious "Missing name" errors after the real data ends.
        if all(v is None or str(v).strip() == "" for v in raw_row):
            continue

        row = [str(v).strip() if v is not None else "" for v in raw_row]

        def get_val(field):
            idx = col_map.get(field)
            if idx is None or idx < 0 or idx >= len(raw_row):
                return ""
            return row[idx]

        def get_raw(field):
            idx = col_map.get(field)
            if idx is None or idx < 0 or idx >= len(raw_row):
                return None
            return raw_row[idx]

        first_name = get_val("first_name")
        last_name = get_val("last_name")

        if not first_name or not last_name:
            errors.append({"row": row_num,
                           "reason": "Missing First Name or Last Name."})
            continue

        # Phone — canonical Indian mobile validation, same as send_sms uses.
        try:
            canonical = normalize_indian_phone(get_val("phone"))
            phone_10 = canonical[3:]  # drop the +91 prefix; DB stores 10 digits
        except ValueError as e:
            errors.append({"row": row_num, "reason": str(e)})
            continue

        # Duplicate detection — per lodge.
        if db.query(Customer).filter(
            Customer.phone == phone_10,
            Customer.lodge_id == lodge_id,
        ).first():
            duplicates += 1
            continue

        # Optional fields
        email = get_val("email") or None
        address = get_val("address") or None
        city = get_val("city") or None
        state = get_val("state") or None
        nationality = get_val("nationality") or "Indian"

        # ID type / number. Type defaults to 'aadhar' (the most common in
        # India) only if id_number is supplied — never invent a number.
        raw_id_type = (get_val("id_type") or "").lower()
        id_type = ID_TYPE_MAP.get(raw_id_type, "aadhar" if get_val("id_number") else "aadhar")
        id_number = get_val("id_number") or None

        if not id_number:
            # id_number is NOT NULL on the model; we record a placeholder
            # 'IMPORTED' marker so operators can find these rows later and
            # collect the real ID at the front desk before check-in.
            id_number = "IMPORTED"

        # Date of birth — accept Excel datetimes or string forms.
        try:
            dob = _parse_dob(get_raw("date_of_birth"))
        except ValueError as e:
            errors.append({"row": row_num, "reason": str(e)})
            continue

        # Gender — Male/M -> M, Female/F -> F, anything else -> Other.
        normalized_gender = None
        raw_gender = (get_val("gender") or "").strip().upper()
        if raw_gender.startswith("M"):
            normalized_gender = "M"
        elif raw_gender.startswith("F"):
            normalized_gender = "F"
        elif raw_gender.startswith("O"):
            normalized_gender = "Other"

        try:
            customer = Customer(
                lodge_id=lodge_id,
                first_name=first_name.title(),
                last_name=last_name.title(),
                phone=phone_10,
                email=email,
                address=address,
                city=city,
                state=state,
                id_type=id_type,
                id_number=id_number,
                date_of_birth=dob,
                nationality=nationality,
                gender=normalized_gender,
                imported_from_excel=True,
            )
            db.add(customer)
            db.flush()
            imported += 1
        except Exception as e:
            db.rollback()
            errors.append({"row": row_num, "reason": f"DB insert failed: {e}"})
            continue

    db.commit()

    return {
        "total_processed": len(rows) - 1,
        "imported": imported,
        "duplicates_skipped": duplicates,
        "errors": len(errors),
        "error_details": errors[:100],
        "message": (f"Import complete: {imported} imported, "
                    f"{duplicates} duplicates skipped, {len(errors)} errors")
    }
